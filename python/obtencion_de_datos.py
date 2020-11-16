import pandas as pd
import openpyxl


def load_data(name:str):
    wb = openpyxl.load_workbook(name, read_only=True, data_only=True)
    ws_capacidad_instalada = wb[wb.sheetnames[0]]
    ws_restricciones_operativas = wb[wb.sheetnames[1]]
    ws__termicas = wb[wb.sheetnames[2]]
    ws_hidros = wb[wb.sheetnames[3]]

    capacidad_instalada = list(ws_capacidad_instalada.values)
    restricciones_operativas = list(ws_restricciones_operativas.values)
    datos_de_termicas = list(ws__termicas.values)
    datos_de_hidros = list(ws_hidros.values)
    wb.close()
    return (capacidad_instalada, restricciones_operativas, datos_de_termicas, datos_de_hidros)

datos = load_data('./excel/cnd_data.xlsx')

# Crear Dataframes
df_cap_instalada = pd.DataFrame(datos[0][1:], columns=[name.rstrip() for name in datos[0][0]])
df_restricciones = pd.DataFrame(datos[1][2:], columns=[name.rstrip() for name in datos[1][1]])
df_termicas = pd.DataFrame(datos[2][1:], columns=[name.rstrip() for name in datos[2][0]])
df_hidros = pd.DataFrame(datos[3][1:], columns=[name.rstrip() for name in datos[3][0]])
dataframes = [df_cap_instalada, df_restricciones, df_termicas, df_hidros]
new_data = []
for data in dataframes:
    new_data.append(data.loc[:,~data.columns.duplicated()])
dataframes = tuple(new_data)


# Cargar la demanda eléctrica
wb = openpyxl.load_workbook('./excel/demanda.xlsx', read_only=True, data_only=True)
ws = list(wb[wb.sheetnames[0]].values)
demand = pd.DataFrame(ws[4:], columns=[name.rstrip() for name in ws[3]])
wb.close()
